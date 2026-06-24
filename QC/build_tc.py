# -*- coding: utf-8 -*-
"""프리즘 BP 4.2 테스트케이스 워크북 빌더"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

KFONT = "맑은 고딕"
def F(**kw): kw.setdefault("name", KFONT); return Font(**kw)

# ---- 색상 ----
C_TCDEF   = "1F4E79"  # TC 정의 헤더(남색)
C_QC      = "2E7D32"  # QC 헤더(녹색)
C_DEFECT  = "B5482A"  # 결함/개발 헤더(주황적)
C_ETC     = "595959"  # 비고
FILL_PASS = "E2EFDA"; FILL_FAIL = "FCE4E4"; FILL_BLOCK = "EDEDED"; FILL_DONE = "E2EFDA"
FILL_P1   = "FDE9E9"
FILL_BAND = "F5F8FB"

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["담당","TC ID","대분류","중분류","우선\n순위","테스트 시나리오(케이스명)","사전조건",
           "테스트 절차","입력/조건","기대결과",
           "QC결과","실행일","QC담당자",
           "결함 내용","심각도","개발 수정상태","재확인(QC)","비고"]
# 그룹별 헤더색 (맨 앞 '담당' 포함 TC정의 10칸)
HCOLORS = [C_TCDEF]*10 + [C_QC]*3 + [C_DEFECT]*4 + [C_ETC]
WIDTHS  = [10,10,14,16,7,30,22,30,18,40, 9,11,10, 28,9,13,12,16]

# 컬럼 인덱스(1-base) — 맨 앞 '담당' 추가로 TC ID 이후 전부 +1
COL_TCID=2; COL_PRIO=5; COL_QC=11; COL_SEV=15; COL_FIX=16; COL_RECHK=17

PROC = {"D":"매니저 화면 분석값과 결과지 노출값을 글자·자릿수 단위로 대조",
        "R":"기대결과의 규칙대로 기대값을 재계산한 뒤 결과지와 대조",
        "S":"HTML 텍스트·구조를 regex/DOM 카운트로 검사",
        "M":"헤드리스 브라우저로 렌더 후 크기·bbox·overflow 측정",
        "E":"스크린샷으로 육안 확인(필요 시 diff)"}
TAGLBL = {"D":"[데이터대조]","R":"[규칙계산]","S":"[정적검사]","M":"[렌더측정]","E":"[육안]"}

# ====================== 결과지 105 ======================
# (중분류, 우선순위, 케이스명, 기대결과, tag)
RPT = [
("A. 빌드 무결성","P1","A1 미치환 토큰 0건","HTML 본문에 {{...}} 패턴(미치환 토큰)이 1건도 남아있지 않다. regex \\{\\{[A-Z_]+\\}\\} → 0건","S"),
("A. 빌드 무결성","P2","A2 다운로드 PDF 파일명 규칙","프리즘워크에서 다운로드한 PDF 파일명이 정해진 규칙([제품]-[지원자명]-[날짜])과 일치","S"),
("B. 메타·식별","P1","B1 채용공고명 일치","meta.channel_title이 doc-header(.doc-header-title, 전 페이지)와 블루헤더(.bp-header-title) 두 위치에서 동일","D"),
("B. 메타·식별","P1","B2 지원자명 일치","meta.candidate가 doc-header(.doc-header-name)·카드(.bp-prof-name)·<title> 3위치에서 동일","D"),
("B. 메타·식별","P1","B3 응시번호 일치","meta.candidate_number가 doc-header(.doc-header-num)·카드(.bp-prof-num) 일치","D"),
("B. 메타·식별","P2","B4 직무명","meta.position이 카드 .bp-prof-job와 일치","D"),
("B. 메타·식별","P2","B5 평가일","meta.evaluation_date가 카드 .bp-date와 일치('평가일: ' 접두 포함)","D"),
("B. 메타·식별","P1","B6 doc-header 전 페이지 동일","모든 페이지 doc-header의 B1~B3 값이 동일(한 페이지라도 다르면 실패)","S"),
("C. 점수 정확성","P1","C1 종합점수 4곳 동일","scores.total이 카드 .bp-score-num·요약 .summary-pos-score·요약막대 .cbar-fill width(%)에서 동일","D"),
("C. 점수 정확성","P1","C2 직무적합도","scores.직무적합도가 P1 .jma .metric-score·P4+ dt-cat '직무적합도 …점'에서 일치","D"),
("C. 점수 정확성","P1","C3 조직적합도","scores.조직적합도가 P1 조직카드 .metric-score·P4+ dt-cat에서 일치","D"),
("C. 점수 정확성","P1","C4 지원동기","scores.지원동기가 P1 지원동기카드·P4+ dt-cat에서 일치","D"),
("C. 점수 정확성","P1","C5 응답완성도 3종","scores.답변적합도/구체성/본인소개가 P1 사이드박스 .side-score·P4+ dt-cat에서 일치","D"),
("C. 점수 정확성","P1","C6 소수 2자리 표기","모든 점수가 \\d+\\.\\d{2} 형식(정수여도 68.00). 1·3자리 노출 시 실패","S"),
("C. 점수 정확성","P1","C7 직무적합도 비교","(scores.직무적합도 − benchmarks.avg.직무적합도) 재계산값과 부호·수치·클래스 일치","R"),
("C. 점수 정확성","P1","C8 조직적합도 비교","(scores.조직적합도 − benchmarks.avg.조직적합도) 일치","R"),
("C. 점수 정확성","P1","C9 지원동기 비교","(scores.지원동기 − benchmarks.avg.지원동기) 일치","R"),
("C. 점수 정확성","P1","C10 부호·색상 분기","차이≥0 → '+'/.up(파랑), 음수 → '−'(U+2212)/.down(주황). 0은 +/.up. 하이픈(-) 아닌 마이너스(−) 확인","R"),
("C. 점수 정확성","P1","C11 요약 종합막대","fill width=total, .dot-avg left=avg.total, .dot-top left=top10.total. 합불 모드 시 .dot-cut left=cut.score","R"),
("C. 점수 정확성","P1","C12 미니막대","fill width=메트릭 점수, .dot-avg=avg.해당, .dot-top=top10.해당","R"),
("C. 점수 정확성","P2","C13 막대 범례 수치","'평균(점수)'=avg.*, '상위 10%(점수)'=top10.* (2자리). 합불 모드 시 컷 범례=cut.label+cut.score","D"),
("C. 점수 정확성","P2","C14 위치점 범위","위치 점 left가 0~100% 범위를 벗어나지 않음(음수/100 초과 없음)","S"),
("D. 등급 정확성","P1","D1 직무적합도 factor 등급","evaluation.직무적합도 중 level=factor 3개(직무경험·직무지식·직무동기)가 .gc-left .grow에 순서·등급 일치","D"),
("D. 등급 정확성","P1","D2 직무역량 소제목 등급","level=subcat-header 항목이 .gc-right-head에 표시","D"),
("D. 등급 정확성","P1","D3 직무역량 하위 8항목 등급","level=sub 8항목(전략적사고~책임감)이 .srow에 순서·등급 일치(홀수 시 빈칸 행 1개 자동)","D"),
("D. 등급 정확성","P1","D4 조직적합도 항목 등급","evaluation.조직적합도.항목 전체가 .wrow에 일치","D"),
("D. 등급 정확성","P1","D5 지원동기 항목 등급","evaluation.지원동기.항목 전체가 지원동기 카드 .grow에 일치","D"),
("D. 등급 정확성","P1","D6 칩 글자·색 동시 일치","상→상(파랑)/중→중(회색)/하→하/없음→'-'(g-none). 글자와 색 클래스가 동시에 일치","R"),
("D. 등급 정확성","P1","D7 등급 항목 개수","JSON 항목 배열 길이 = 렌더 행 수(빈칸 보정행 제외), 누락·중복 없음","D"),
("E. 합불 배지","P1","E1 직군 위치 배지","항상 출력. 텍스트='{benchmarks.group} 상위 {benchmarks.percentile}%'","D"),
("E. 합불 배지","P1","E2 합불 배지 출력 여부","meta.pass_badge.mode=null이면 .bp-badge.fill·.dot-cut·컷 범례 모두 미출력, null 아니면 출력","R"),
("E. 합불 배지","P1","E3 합불 배지 텍스트","meta.pass_badge.label과 일치","D"),
("E. 합불 배지","P1","E4 합불 배지 색 분기","'이내'/'권장'→.fill(파랑 #e0f3ff), '초과'/'검토 필요'→.fill.over(주황 #d75d00)","R"),
("E. 합불 배지","P1","E5 컷 막대 동반","합불 모드일 때만 요약막대에 컷 점(.dot-cut)+컷 범례(cut.label+cut.score) 동반","R"),
("E. 합불 배지","P2","E6 합불 배지 정책 불변·전 라이선스 적용","합불 배지(적격/검토 필요) 사용 여부는 모든 라이선스에 적용. 4.2에서 정책·기능 변경 없음(기존 로직 유지, 모든 경우의 수를 다 표현하진 않음)","R"),
("F. 검증 포인트","P2","F1 강점 카드","verification.strengths와 개수·순서 일치, 배지 클래스 .vp-badge.up","D"),
("F. 검증 포인트","P2","F2 검토 카드","verification.reviews와 개수·순서 일치, 배지 클래스 .vp-badge.down","D"),
("F. 검증 포인트","P2","F3 factor명·등급칩",".nm=factor, .vp-grade=등급 매핑(상/중/하/-)","D"),
("F. 검증 포인트","P2","F4 대분류 라벨 매핑",".cat=FACTOR_AREA[factor] 매핑값 일치(매핑 없으면 .cat 미표시)","R"),
("F. 검증 포인트","P2","F5 백분위 배지",".vp-badge pct 텍스트 일치(예 '상위 18%'·'하위 12%')","D"),
("F. 검증 포인트","P2","F6 요약·면접연계 문구",".vp-sum=comment, .vp-tip .ttxt=interview_question","D"),
("F. 검증 포인트","P2","F7 하단 하드스킬 칩","강점카드 하단=확인된 하드스킬(검출), 검토카드 하단=확인되지 않은 하드스킬(미검출)","R"),
("G. 문항·자소서","P1","G1 문항 테이블 행 수","essay.questions 개수와 일치, 각 행 '문항 {no}'+title","D"),
("G. 문항·자소서","P1","G2 답변적합도 값",".qmetric .v = question.answer_fit(2자리)","D"),
("G. 문항·자소서","P1","G3 AI 작성 의심 표기",".pct=question.gpk(탐지/미탐지). 탐지→주황(#f09000), 미탐지→.nd 회색. ai_rate(%)는 미노출","R"),
("G. 문항·자소서","P1","G4 P3 카드 수","essay.questions 개수와 일치, 헤드 '문항 {no}'+title","D"),
("G. 문항·자소서","P1","G5 본문=sentences 전체",".p3-body에 모든 sentences[].text가 누락 없이 순서대로 포함","D"),
("G. 문항·자소서","P1","G6 근거문장 강조",".hl이 sentences[].hl==true 문장만 감쌈, false 문장엔 강조 없음","R"),
("G. 문항·자소서","P2","G7 P3 본문 gpk 밑줄 제거","P3 본문에 문장단위 gpk 밑줄 없음(문항 단위 .p3-ai-val만 표기)","S"),
("G. 문항·자소서","P1","G8 메트릭 칩",".p3-metric-group=detected.metrics(name+등급칩), 개수·순서 일치","D"),
("G. 문항·자소서","P1","G9 역량 태그",".p3-skill-tag=detected.factors, 개수·순서 일치(쉼표 구분)","D"),
("G. 문항·자소서","P1","G10 문항 AI 작성 의심",".p3-ai-val=gpk(탐지/미탐지), 색 분기 G3 동일","R"),
("H. 응답완성도","P2","H1 3박스 존재","답변적합도·구체성·본인소개 순서로 .side-box 3개","S"),
("H. 응답완성도","P1","H2 점수=scores",".side-score=scores.*(2자리)","D"),
("H. 응답완성도","P2","H3 코멘트=short_text",".side-cmt=completeness.*.short_text 문구 일치","D"),
("H. 응답완성도","P1","H4 주의 배지 분기","점수<40 항목에만 .warn-badge(주의) 출력, ≥40은 미출력","R"),
("I. 레이더","P2","I1 축 라벨 6개·순서","RADAR_LABELS=직무적합도/조직적합도/지원동기/답변완성도/구체성/자기소개 순서, 각 라벨에 점수(2자리) 동반","R"),
("I. 레이더","P2","I2 지원자 데이터","RADAR_USER=[직무적합도,조직적합도,지원동기,답변적합도,구체성,본인소개] 점수(2자리)","R"),
("I. 레이더","P2","I3 평균 데이터","RADAR_AVG=benchmarks.radar_avg 배열(원본값 그대로)","D"),
("I. 레이더","P2","I4 차트 렌더","<canvas id='radarChart'>에 Chart.js 렌더, JS 콘솔 에러 없음","M"),
("J. BP 평가 상세","P1","J1 3개 대분류","dt-cat 직무적합도·조직적합도·지원동기, 각 점수+'점'","D"),
("J. BP 평가 상세","P1","J2 항목별 코멘트",".dt-comment=항목 summary. summary 비면 '해당 역량을 확인할 수 없음' 출력","R"),
("J. BP 평가 상세","P1","J3 근거문장=evidence_ids",".dt-list li가 evidence_sentences[id] 문장과 일치(id 순서대로)","D"),
("J. BP 평가 상세","P1","J4 근거 없음 처리","evidence_ids 빈 배열이면 .dt-empty '확인된 자기소개서 근거 문장이 없습니다.' 출력","R"),
("J. BP 평가 상세","P1","J5 등급 배지",".dt-badge=항목 grade 매핑(상/중/하/-), subcat-header도 배지 표시","D"),
("J. BP 평가 상세","P2","J6 factor/sub 구분","level=factor는 .nm, 하위(sub)는 .dt-tag","R"),
("J. BP 평가 상세","P2","J7 완성도 3대분류+코멘트행","dt-cat(점수)+dt-commentrow 구성","S"),
("J. BP 평가 상세","P1","J8 룰베이스 문장 구간 일치","rule_text(metric,score) 재계산과 .dt-commentrow 글자 단위 일치(경계값 79/80, 59/60 주의)","R"),
("J. BP 평가 상세","P2","J9 하드스킬 대분류","dt-hardcat '하드스킬'(네이비 #495e76 배경)","S"),
("J. BP 평가 상세","P1","J10 하드스킬 행=hard_skills","hard_skills.항목과 개수·순서·스킬명 일치","D"),
("J. BP 평가 상세","P1","J11 검출/미검출 표기","status=검출→.dt-detect.yes '검출'(파랑)+근거리스트, 미검출→.dt-detect.no '미검출'(주황)+'관련 근거 문장 없음.'","R"),
("J. BP 평가 상세","P1","J12 검출 항목 근거문장","evidence_ids 해석값이 .dt-list에 일치","D"),
("K. 누락·품질","P1","K1 evidence_ids 참조 무결","모든 evidence_ids가 evidence_sentences에 존재(미정의 id 0건)","D"),
("K. 누락·품질","P1","K2 빈 항목 일관 처리","grade 없음+빈 summary 항목이 코멘트 fallback+근거 empty로 일관 표기","R"),
("K. 누락·품질","P1","K3 길이 제약","summary.headline ≤54자, 각 bullet ≤60자, 불릿 ≤10개(초과 시 빌드 실패·P1 넘침)","S"),
("K. 누락·품질","P2","K4 불릿 권장 개수","summary.bullets 응시자별 4~5개 권장(5 초과 시 P1 넘침 위험)","S"),
("K. 누락·품질","P1","K5 응시자 간 값 혼입 없음","A응시자 HTML에 B응시자 이름/번호/점수가 섞이지 않음","D"),
("L. A4 규격","P1","L1 .page 990×1399 overflow hidden","모든 .page가 width:990px·height:1399px·overflow:hidden. min-height 사용 금지","S"),
("L. A4 규격","P1","L2 @page A4","@page{size:A4;margin:0} + print-color-adjust:exact 포함","S"),
("L. A4 규격","P1","L3 모든 페이지 doc-header","모든 페이지에 doc-header(27px), 자동생성 P4+ 포함 예외 없음","S"),
("L. A4 규격","P1","L4 모든 페이지 footer","모든 페이지에 footer(34px), 예외 없음","S"),
("L. A4 규격","P1","L5 P1에만 지원자 카드",".bp-card는 1페이지에 1개, 2페이지 이후 없음","S"),
("M. 푸터·고정문구","P2","M1 푸터 로고 경로","../logo/logo_prism.png","S"),
("M. 푸터·고정문구","P2","M2 푸터 제품명","'서류평가'","S"),
("M. 푸터·고정문구","P1","M3 저작권 2줄 정확 일치","'Copyright © 2026 muhayu Inc. All rights reserved.' + '본 결과의 평가 기준, 문항, 분석 내용 등 모든 지적재산권은 ㈜무하유에 귀속됩니다. 무단 복제 및 재배포를 금합니다.'","S"),
("M. 푸터·고정문구","P1","M4 페이지번호 동적·연속",".footer-page가 n/total 형식, total=실제 .page 수, 1부터 빠짐없이 증가","M"),
("M. 푸터·고정문구","P1","M5 마지막 페이지 안내 2줄",".report-info 2개 item(1=AI 분석 고지 프리즘 BP 문구, 2=개인정보 법령 고지), CLAUDE.md 원문과 글자 단위 일치","S"),
("M. 푸터·고정문구","P1","M6 마지막장 저작권·AI기본법 고지 필수","모든 결과지 마지막 장 끝에 저작권 및 AI기본법 관련 고지가 반드시 포함. 분량으로 페이지가 추가되더라도 반드시 추가(누락 시 치명)","S"),
("N. 페이지네이션","P2","N1 #dt-source 제거","숨김 소스가 렌더 후 제거되어 화면/인쇄에 중복 노출 안 됨","M"),
("N. 페이지네이션","P2","N2 섹션 헤더 고아 없음","dt-cat/dt-subcat/dt-hardcat/dt-sec-title가 페이지 끝에 혼자 남지 않고 다음 블록과 동반","M"),
("N. 페이지네이션","P2","N3 연속 페이지 첫 행 구분선","2번째 상세페이지부터 첫 콘텐츠 행에 .dt-conttop(상단 구분선)","S"),
("N. 페이지네이션","P1","N4 블록 중간 잘림 없음","어떤 .blk도 페이지 경계에서 분리되지 않음(한 블록=한 페이지 내)","M"),
("O. 오버플로·레이아웃","P1","O1 콘텐츠 넘침 없음","각 .page 자식 콘텐츠가 1399px 초과 안 함(최하단 요소 bottom ≤ 페이지 bottom)","M"),
("O. 오버플로·레이아웃","P1","O2 P1 고정 영역 미초과","헤더+카드+요약이 정해진 높이를 넘지 않아 하단 2단 콘텐츠가 잘리지 않음","M"),
("O. 오버플로·레이아웃","P1","O3 고정 높이 카드 유지",".jma·.chart-card 444px 고정(눌리거나 늘어나지 않음)","M"),
("O. 오버플로·레이아웃","P1","O4 텍스트 잘림 없음","헤드라인·문항 제목·근거문장·코멘트가 …/클리핑 없이 전체 노출","M"),
("O. 오버플로·레이아웃","P3","O5 막대·점·칩 영역 내","cbar 점이 트랙 밖으로, 등급칩이 셀 밖으로 튀지 않음","E"),
("O. 오버플로·레이아웃","P3","O6 2열 그리드 정렬",".srow/.wrow 2열이 깨지지 않고 라벨-칩이 같은 기준선 정렬","E"),
("O. 오버플로·레이아웃","P3","O7 빈 칸 보정행 자연스러움","직무역량 홀수 항목 시 추가된 빈 행에 잔여 테두리/이상 여백 없음","E"),
("P. 폰트·인코딩·표기","P2","P1 폰트 로드 정상","Noto Sans KR/Noto Sans <link> 정상, 숫자에 .num(Noto Sans) 적용","M"),
("P. 폰트·인코딩·표기","P1","P2 한글 깨짐 없음","<meta charset='UTF-8'>, 모지바케·물음표박스 0건","S"),
("P. 폰트·인코딩·표기","P3","P3 글자 깨짐(tofu) 없음","특수문자(℃·㈜·©·−··) 정상 렌더","E"),
("P. 폰트·인코딩·표기","P1","P4 띄어쓰기/오탈자","근거문장·자소서 본문의 어절 중간 공백 오류 점검(예 '반 복한','데 이터'). 한글 1~2음절 뒤 공백+한글 비정상 패턴 스캔","S"),
("P. 폰트·인코딩·표기","P2","P5 색상 토큰 일치","등급칩·배지·강조색이 BP 토큰과 일치(높음 rgba(0,117,255,0.6)/검토 #d75d00/근거강조 rgba(0,117,255,0.1)/AI의심 #f09000)","S"),
("P. 폰트·인코딩·표기","P3","P6 강조 요소 과다 없음","한 페이지 Bold 32px↑ 2개 이내, 색상 강조 3종 이내","E"),
("Q. 라이선스별 구성","P1","Q1 단독: 종합결과 페이지 없음","라이선스 단독 사용 시 종합결과 페이지를 사용하지 않음","S"),
("Q. 라이선스별 구성","P1","Q2 단독: 결과지 상단 응시자 정보","라이선스 단독 사용 시 결과지 상단에 응시자 정보 영역을 표현","S"),
("Q. 라이선스별 구성","P1","Q3 복수(2개+): 종합결과 페이지 사용","라이선스 2개 이상 사용 시 종합결과 페이지를 사용","S"),
("Q. 라이선스별 구성","P1","Q4 복수: 종합결과 상단 응시자 정보","종합결과 페이지 상단에 응시자 정보 영역을 표현","S"),
("Q. 라이선스별 구성","P1","Q5 복수: 종합결과 구성(판정+각 라이선스 요약)","종합결과 페이지가 '종합결과 판정 + 각 라이선스 요약 결과'로 구성","S"),
("Q. 라이선스별 구성","P1","Q6 복수: 종합결과 1페이지 이내","종합결과 페이지가 1페이지를 넘기지 않음","M"),
("Q. 라이선스별 구성","P1","Q7 복수: 각 라이선스 결과지 상단 응시자 정보 없음","라이선스 2개 이상일 때 각 라이선스 결과지 상단에는 응시자 정보 영역을 표현하지 않음(응시자 정보는 종합결과 페이지에만)","S"),
("Q. 라이선스별 구성","P2","Q8 단독/복수 표기 차이","종합점수·상위n%·전형 통과 여부 표기가 BP 단독/복수 라이선스에 따라 다름","R"),
("R. 평가요약·지원자현황","P2","R1 평가 요약 영역 구성","평가 요약 영역에 한 줄 코멘트·자소서 특장점·지원자 현황 막대그래프가 출력","S"),
("R. 평가요약·지원자현황","P2","R2 지원자 현황 막대(위치막대)","지원자 현황 막대 = 종합점수/메트릭 아래 위치막대(cbar)로 표현(figma 확인 2026-06-24). C11~C13(막대 위치·범례)으로 검증","R"),
("S. 전형 통과 여부","P1","S1 전형 통과 여부 텍스트","'적격/부적격' 대신 '전형 통과 기준'(종합점수/배수 기준) 텍스트로 표기(AI기본법 표현 준수)","S"),
("S. 전형 통과 여부","P1","S2 상위 n% 표기","BP평가 점수에 직군 내 상위 n%가 동반 표기","D"),
("T. GPK 분석·과금","P2","T1 자소서 분석 GPK 기본 노출","자기소개서 분석 내용에 GPK 결과(AI 작성률·AI 작성 의심 문장)가 기본 제공","S"),
("T. GPK 분석·과금","P2","T2 GPK 미선택 고객사 과금 미반영","GPK 라이선스 미선택 고객사도 분석 시 GPK 실행하되 수불부·사용량에 미집계(백엔드/과금 검증)","R"),
("U. 엑셀 결과지","P1","U1 전형 통과 여부 텍스트","엑셀파일 결과지에 변경된 전형 통과 여부 텍스트가 반영","D"),
("U. 엑셀 결과지","P1","U2 하드스킬 검출 여부 추가","엑셀파일 결과지에 하드스킬 검출 여부가 추가 표기","D"),
("U. 엑셀 결과지","P2","U3 점수·등급 원본 일치","엑셀 결과지 값이 원본 데이터와 일치(누락·혼입 없음)","D"),
("V. 타 라이선스 페이지","P2","V1 비-BP 검사 결과 페이지 렌더 정상","결함·표절·블라인드·GPK 등 BP 외 검사 결과 페이지가 깨짐·잘림·빈 페이지 없이 정상 렌더(4.2에서 콘텐츠 미변경, family 디자인 표현만 적용)","E"),
("V. 타 라이선스 페이지","P2","V2 합불 표현 워딩 통일","결과지 전반에서 합불 표현이 '권장/검토 필요'(전형 통과 기준)로 표기되고 '적격/부적격' 직접표현이 어느 페이지에도 없음","S"),
]

def rpt_rows():
    rows=[]
    for sub,prio,name,exp,tag in RPT:
        inp = "3단계 분석값(매니저 결과확인 화면)" if tag in ("D","R") else "-"
        pre = "샘플 자소서가 접수→분석까지 완료 · 엑셀 결과지 생성됨" if sub.startswith("U.") else "샘플 자소서가 접수→분석→결과지까지 완료(매니저 화면에 분석값 확인 가능)"
        rows.append(["결과지(출력물)",sub,prio,name,pre,PROC[tag],inp,exp,TAGLBL[tag]])
    return rows

# ====================== API 42 ======================
# (중분류, 우선순위, 케이스명, 사전조건, 절차, 입력/조건, 기대결과, 비고)
PRE_OL = "execute API·공방 프롬프트(bp-v4-comment) 배포 · API PRISM/GENERATE-BP-COMMENT · llmServiceKey 발급"
PRE_HL = "execute API·공방 프롬프트(bp-point-summary) 배포"
PRE_HX = "하드스킬 추출 프롬프트(공방 extract-jd-hardskill) 배포 · JD 입력 가능"
PRE_HD = "COMMON/EXTRACT-EVIDENCE(공방 hardskill-extract-evidence-test, 폴백 EVALUATE) 가용 · 1단계 하드스킬+검출기준 확보"
API = [
("① 한 줄 코멘트","P1","정상 생성(happy path)",PRE_OL,"POST /common/prompt-execute 호출","강점2+검토1~2 factor digest, max_chars=\"90\"","status=ok, result.one_liner 1문장, parts.strength/weakness/action 3요소 모두 채워짐, meta.grounded=true",""),
("① 한 줄 코멘트","P1","두괄식(강점 선두)",PRE_OL,"응답 one_liner 문장 구조 확인","정상 digest","한 문장이 강점→단점→액션 순서, 강점이 문장 맨 앞",""),
("① 한 줄 코멘트","P1","금지어 미사용",PRE_OL,"one_liner·parts에 금지어 포함 검사","정상 digest","'합격/불합격/적격/부적격/탈락' 미포함('권장/검토 필요/확인'으로 표현)","hard fail"),
("① 한 줄 코멘트","P1","무근거 금지/grounded",PRE_OL,"입력에 없는 영역·등급 등장 검사","grade=none factor 포함 digest","none 역량을 강점으로 단정 안 함, 입력에 없는 사실 미생성, meta.grounded=true","hard fail"),
("① 한 줄 코멘트","P1","길이 제한",PRE_OL,"one_liner 실제 글자수 확인","max_chars=90","one_liner 글자수 ≤ 90(초과 시 비교수치→수식어 순 축소)",""),
("① 한 줄 코멘트","P2","3요소 누락 시 재생성",PRE_OL,"parts 각 필드 확인","강점/검토 일부만 있는 digest","parts.strength/weakness/action 중 빈 값 없음",""),
("① 한 줄 코멘트","P2","검토 메트릭 롤업",PRE_OL,"digest factors의 level 확인","같은 메트릭 none factor 2개+","검토가 level=metric 1개로 롤업('(메트릭) 전반 근거 확인 안 됨' 서술)",""),
("① 한 줄 코멘트","P1","standing 수치 미출력",PRE_OL,"one_liner에 숫자 단정 여부 확인","total_score·percentile_top 포함","상위%·총점 수치가 문장에 단정 출력되지 않음(어조 판단에만 사용)",""),
("① 한 줄 코멘트","P2","강점이 high 아닐 때 완화",PRE_OL,"one_liner 서술 톤 확인","상위2 factor grade=mid","'상대적으로 두드러진' 식 완화 서술(단정 강점 아님)",""),
("① 한 줄 코멘트","P2","출력 스키마 Strict",PRE_OL,"result 구조 검증","정상 호출","required(one_liner·char_count·parts·meta) 전부 존재, additionalProperties 없음",""),
("① 한 줄 코멘트","P2","변수 누락 감지",PRE_OL,"config.onMissingVariables=error로 호출","keyValue에서 max_chars 누락","status=fail + 변수 누락 message(빈 문자열 대체 아님)",""),
("① 한 줄 코멘트","P3","char_count 부정확 주의",PRE_OL,"char_count와 실제 글자수 비교","정상 호출","LLM 자체 계산이라 차이 가능 → 정확값 필요 시 one_liner로 재계산(알려진 한계)","참고"),
("④ 자소서 특장점","P1","정상 요약",PRE_HL,"keyValue.body=자소서 본문으로 호출","자소서 본문","status=ok, result.highlights 배열 반환",""),
("④ 자소서 특장점","P1","개수 4~5개",PRE_HL,"highlights 길이 확인","정상 body","highlights 4~5개(3개 이하/6개 이상 실패)",""),
("④ 자소서 특장점","P2","형식: 20자 안팎 명사구",PRE_HL,"각 항목 길이·종결형 확인","정상 body","각 항목 18~25자 짧은 명사구, 명사·체언 종결('~했다'·'제목: 설명' 아님)",""),
("④ 자소서 특장점","P1","가치판단 금지",PRE_HL,"평가어 포함 검사","정상 body","'좋다/나쁘다·강점/약점·우수/부족' 등 평가어 없음(중립 요약)","hard fail"),
("④ 자소서 특장점","P1","무근거 금지",PRE_HL,"본문 대조","정상 body","본문에 없는 경력·수치·사실 생성 없음, 모든 항목 본문 근거 기반","hard fail"),
("④ 자소서 특장점","P2","중복 금지",PRE_HL,"항목 간 의미 중복 확인","정상 body","같은 내용 반복 항목 없음",""),
("④ 자소서 특장점","P1","금지어 미사용",PRE_HL,"금지어 검사","정상 body","합격/불합격/적격/부적격/탈락 미포함","hard fail"),
("④ 자소서 특장점","P2","마크업·번호 없음",PRE_HL,"출력 문자열 검사","정상 body","#·##·**·- 마크업, ①②③·1. 번호 없이 순수 문장",""),
("④ 자소서 특장점","P2","입력=body만",PRE_HL,"입력 구성 확인","head+body 동시 전달 시도","질문 head 미포함, 본문만으로 요약(여러 문항이면 답변 이어붙임)",""),
("④ 자소서 특장점","P3","결과지 매핑",PRE_HL,"P1 summary-bullets 매핑 확인","정상 highlights","highlights[]가 결과지 P1 요약카드 불릿에 그대로 매핑(소비측 분해 불필요)",""),
("② 하드스킬 추출","P1","정상 추출",PRE_HX,"keyValue.jd_text=JD plain text로 호출","JD 본문","status=ok, required_hard_skills 5~7개 반환",""),
("② 하드스킬 추출","P1","개수 5~7개",PRE_HX,"목록 길이 확인","정상 JD","required_hard_skills 5~7개(미만/초과 재실행)",""),
("② 하드스킬 추출","P1","3범주만",PRE_HX,"category 값 확인","정상 JD","category가 tool_based/method_based/knowledge_based만(경험·산출물 유형 추출 안 함)",""),
("② 하드스킬 추출","P1","evidence_from_input 원문 보존",PRE_HX,"원문 대조","정상 JD","각 항목 evidence_from_input에 JD 원문 일부 그대로 보존(무근거 금지)",""),
("② 하드스킬 추출","P1","검출 기준 동반 생성",PRE_HX,"detection_criteria 확인","정상 JD","각 하드스킬에 detection_criteria(ⓐ인정+ⓑ배제 조건) 존재, 누락 시 보강",""),
("② 하드스킬 추출","P1","Hard↔Soft 혼동 금지",PRE_HX,"결과 검사","'성실'·'소통' 등 태도 포함 JD","소프트스킬(태도·성향) 미추출, 혼입 시 실패","hard fail"),
("② 하드스킬 추출","P1","필수 누락 시 중단",PRE_HX,"응답 확인","job_title·핵심업무 없는 JD","job_title/key_responsibilities 없으면 중단·재요청(required_competencies 없으면 진행하되 하드스킬 빌 수 있음)",""),
("② 하드스킬 추출","P2","임의 추가 금지",PRE_HX,"결과 검사","팀·연차 미기재 JD","팀명·부서·경력연차·학력·보상·근무지·고용형태를 임의 추정·추가 안 함(미입력은 missing_information)",""),
("② 하드스킬 추출","P2","capability 문장화",PRE_HX,"rewritten_as_capability 확인","정상 JD","단순 키워드가 아닌 '목적어+동사+관찰 가능 결과물' 한 문장",""),
("② 하드스킬 추출","P2","필수/우대 분류",PRE_HX,"requirement_type 확인","우대사항 포함 JD","지원자격·필수→must_have, 우대→nice_to_have 정확 분류",""),
("② 하드스킬 추출","P2","입력 흐름(접수 업로드→매니저 plain text)",PRE_HX,"입력 경로 확인","JD 파일→텍스트","고객사가 접수 화면에서 JD/채용공고 파일 업로드 → 운영자가 다운로드해 매니저에 plain text 직접 입력 → 추출 API 입력=plain text. 키워드 확장 없음(AI 프롬프트 매칭)","확정 2026-06-24"),
("③ 하드스킬 검출","P1","정상 검출",PRE_HD,"key·rubric·body로 호출","자소서+하드스킬 목록","status=ok, result[].key별 evidence(근거문장) 반환",""),
("③ 하드스킬 검출","P1","rubric inline 전달",PRE_HD,"rubric 구성 확인","1단계 검출기준","rubric에 모든 key의 검출 기준 포함(누락 key 없음)",""),
("③ 하드스킬 검출","P1","검출/미검출 판정",PRE_HD,"score/confidence 기준 판정 확인","정상 호출","score 기준 검출/미검출 판정(minEvidenceItem=1이라 evidence ≥1, score 낮으면 미검출)",""),
("③ 하드스킬 검출","P1","무근거 금지",PRE_HD,"evidence 본문 대조","정상 호출","자소서에 실제로 있는 문장만 evidence로 반환(없는 근거 생성 없음)","hard fail"),
("③ 하드스킬 검출","P1","배제 조건 적용",PRE_HD,"오검출 확인","피상적·동음이의 포함 자소서","ⓑ배제 조건(피상적 언급·동음이의) 근거 불인정, 단어 일치만으로 채택 안 함 → score 낮게/미검출",""),
("③ 하드스킬 검출","P2","근거 개수 제약",PRE_HD,"evidence 개수 확인","minEvidenceItem=1,max=5","키별 evidence 1~5개",""),
("③ 하드스킬 검출","P2","offset 반환",PRE_HD,"evidence 위치값 확인","정상 호출","evidence[].bodyIndex/beginOffset/endOffset 반환(결과지 하이라이트용)",""),
("③ 하드스킬 검출","P2","결과지 매핑",PRE_HD,"hard_skills.항목 변환 확인","정상 result","result[]→hard_skills.항목[]{name·status·summary·evidence_ids·score} 매핑",""),
("③ 하드스킬 검출","P3","폴백 동작",PRE_HD,"폴백 경로 확인","EXTRACT-EVIDENCE 미가용","미가용 시 EVALUATE-CUSTOM-FACTOR(grade·reasonSegment)로 폴백 동작",""),
]

def api_rows():
    return [["API",sub,prio,name,pre,proc,inp,exp,memo] for sub,prio,name,pre,proc,inp,exp,memo in API]

# ====================== 어드민 27 ======================
PRE_ADM = "어드민 로그인 · 해당 화면 접근"
NV = "화면정의서 기준 확정 필요"
ADM = [
("검사요청서","P1","벌크형/단건형 전환","검사요청서 화면 진입","벌크형↔단건형 토글","-","선택에 따라 입력 폼이 전환되고 이전 입력이 유실되지 않음"),
("검사요청서","P2","파일 추가/삭제","화면 진입","파일 추가 후 목록 확인, 삭제","테스트 파일","파일명·파일크기 표시, 삭제 시 목록에서 제거"),
("검사요청서","P2","최근 검사요청서 불러오기","이전 요청 이력 존재","불러오기 실행","-","이전 검사요청서 내용이 폼에 채워짐"),
("검사요청서","P2","임시저장","입력 일부 작성","임시저장 클릭","부분 입력","입력 내용 저장, 재진입 시 복원"),
("검사요청서","P1","검사요청 제출(필수값 검증)","필수 항목 미입력","검사요청 클릭","필수값 누락","필수값 누락 시 제출 차단·안내, 충족 시 정상 접수"),
("검사요청서","P3","나가기(미저장 경고)","미저장 변경 존재","나가기 클릭","-","미저장 변경 시 경고 노출"),
("자료 생성","P2","자료생성 모달 오픈","검사 요청 존재","서류결과 자료생성 클릭","-","자료 생성 등록 모달이 열림"),
("자료 생성","P2","파일 업로드","모달 오픈","파일 업로드","결과 파일","파일명 표시, 업로드 반영"),
("자료 생성","P2","멀티스텝 이전/다음","모달 오픈","이전/다음 단계 이동","-","단계 이동 시 입력 유지, 자료생성등록 완료"),
("자료 생성","P3","닫기","모달 오픈","닫기/× 클릭","-","모달 닫힘"),
("근거요약 QC","P1","결함검사·항목 표시","근거요약관리 모달 진입","항목 확인","-","메트릭별 결함검사·등급·근거문장이 표시됨"),
("근거요약 QC","P1","재생성","항목 표시","메트릭 재생성 클릭","-","해당 메트릭 코멘트/근거가 재생성되어 갱신"),
("근거요약 QC","P1","등급 저장","등급 수정","등급 저장 클릭","수정 등급","변경 등급이 저장·반영"),
("근거요약 QC","P2","하이라이트 보기/제거","근거문장 표시","하이라이트 토글","-","근거문장 하이라이트가 표시/제거됨"),
("근거요약 QC","P2","선택불용등록","문장 선택","선택불용등록 클릭","불용 문장","선택 문장이 불용 처리되어 근거에서 제외"),
("근거요약 QC","P1","하드스킬 재매칭","하드스킬 행 표시","재매칭/전체 재매칭 클릭","-","검출 여부·근거문장이 재매칭되어 갱신"),
("근거요약 QC","P2","검토(사전)반영·사전보기","항목 표시","검토 반영·사전보기 클릭","-","사전 검토 내용이 반영/미리보기됨"),
("하드스킬 관리","P1","JD 입력·하드스킬 추출","하드스킬 관리 화면","직무기술서 텍스트 입력 후 추출","JD 텍스트","하드스킬 키워드가 추출되어 표시(②추출 API 연동)"),
("하드스킬 관리","P2","저장된 직무기술서 목록","JD 저장 이력","목록 확인","-","아이디·제목·태그형식·하드스킬·작성일 컬럼 표시"),
("하드스킬 관리","P2","복사/삭제","목록 항목 존재","복사·삭제 클릭","-","항목 복사 생성, 삭제 시 목록에서 제거"),
("하드스킬 관리","P3","태그형식 표시","하드스킬 존재","태그형식 보기","-","하드스킬이 태그 형식으로 표시"),
("하드스킬 관리","P3","다크/라이트 모드","화면 진입","어둡게/밝게 전환","-","테마 전환이 정상 적용"),
("검사요청 확인","P1","6개 검사 섹션 표시","검사요청 완료","확인 화면 진입","-","채용정보·결함검사·표절검사·GPT Killer·블라인드·BP평가 6섹션 모두 표시"),
("검사요청 확인","P2","등급(중대/의심/참고) 표시","검사 설정 존재","등급 확인","-","결함/블라인드 항목별 등급(중대·의심·참고)이 설정값대로 표시"),
("검사요청 확인","P2","부적격 기준·제외 문항","검사 설정 존재","항목 확인","-","검사별 부적격 기준·제외 문항이 설정대로 표시"),
("검사요청 확인","P2","PDF 다운로드","확인 화면","PDF 다운로드 클릭","-","검사 요청 내용이 PDF로 다운로드"),
("검사요청 확인","P3","BP평가 버전 표시","BP평가 포함","버전 확인","-","BP평가 버전이 표시"),
("접수-전형 통과 기준","P1","종합점수 기준 전환","프리즘워크/브릿지 접수 화면","전형 통과 기준 설정","-","합불(적격/부적격) 선택 대신 종합점수 기준 '전형 통과 기준'으로 설정"),
("접수-전형 통과 기준","P1","배수 기준 선택","접수 화면","배수기준 선택","-","배수기준 선택 기능으로 통과 기준 설정 가능"),
("접수-전형 통과 기준","P2","BP v2 라디오버튼 제거","접수 화면","BP v2 선택 UI 확인","-","BP v2 선택 라디오버튼이 화면에서 제거됨(모듈은 유지, 기존 고객사 영향 없음)"),
("접수-전형 통과 기준","P2","직무기술서 업로드(RP매칭 공유)","접수 화면","직무기술서 업로드","JD 파일","직무기술서 업로드 기능이 RP매칭과 공유되어 동작"),
("검사요청 확인-직무기술서","P2","직무기술서 표현","검사요청 확인 화면","직무기술서 영역 확인","-","검사요청 확인 화면에 직무기술서가 표현됨"),
("BP평가 설정-직무기술서","P1","직무기술서 설정","BP평가 설정 화면","직무기술서 설정","-","BP평가 설정 화면에 직무기술서 설정 기능(하드스킬 추출용)이 존재"),
("BP평가 설정-직무기술서","P2","직무별 직무기술서 설정","BP평가 설정 화면","직무별 설정","-","직무별로 직무기술서를 설정할 수 있음"),
("매니저-콘텐츠 관리","P1","한 줄 코멘트 관리","BP평가 결과확인 화면","한 줄 코멘트 확인·수정·재생성","-","한 줄 코멘트를 확인·수정·재생성 관리 가능"),
("매니저-콘텐츠 관리","P1","자소서 특장점 관리","BP평가 결과확인 화면","특장점 확인·수정·재생성","-","자소서 특장점을 확인·수정·재생성 관리 가능"),
("매니저-콘텐츠 관리","P1","근거요약 관리","BP평가 결과확인 화면","근거요약 확인·재생성","-","근거요약을 확인·재생성 관리 가능(근거요약 QC와 연계)"),
("매니저-콘텐츠 관리","P1","하드스킬 근거문장 관리","BP평가 결과확인 화면","근거문장 확인·재매칭","-","하드스킬 근거문장을 확인·재매칭 관리 가능"),
("하드스킬 키워드 관리","P1","문장형식 제거→텍스트형식 추가","직무기술서 게시판 편집","편집 컬럼 형식 확인","-","편집 컬럼에서 '문장형식' 제거, 하드스킬용 '텍스트형식' 추가"),
("하드스킬 키워드 관리","P1","키워드 생성/재생성","직무기술서 편집","하드스킬 키워드 생성·재생성","JD 텍스트","하드스킬 키워드 생성 및 재생성 동작"),
("하드스킬 키워드 관리","P2","키워드 삭제/추가","하드스킬 키워드 존재","삭제·수동 추가","-","하드스킬 키워드 삭제 및 수동 추가 가능"),
("프리즘워크-다운로드 설정","P2","BP 콘텐츠 옵션 변경","프리즘워크 다운로드 설정","옵션 확인","-","변경된 BP 콘텐츠에 맞춰 다운로드 옵션이 변경됨"),
("프리즘워크-다운로드 설정","P2","'BP평가 상세(근거 문장)' 명칭","다운로드 설정","옵션명 확인","-","'근거문장' 옵션명이 'BP평가 상세(근거 문장)'로 변경 반영"),
]

def adm_rows():
    return [["어드민",sub,prio,name,pre,proc,inp,exp,NV] for sub,prio,name,pre,proc,inp,exp in ADM]

# ====================== 워크북 ======================
wb = Workbook()

def style_header(ws):
    for c,(h,col) in enumerate(zip(HEADERS,HCOLORS),1):
        cell=ws.cell(1,c,h)
        cell.font=F(bold=True,color="FFFFFF",size=9)
        cell.fill=PatternFill("solid",fgColor=col)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=BORDER
    for c,w in enumerate(WIDTHS,1):
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=30
    ws.freeze_panes="F2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(HEADERS))}1"

def add_dv(ws,n):
    defs=[(COL_QC,'"Pass,Fail,N/A,Blocked"'),(COL_PRIO,'"P1,P2,P3"'),
          (COL_SEV,'"치명,중,경"'),(COL_FIX,'"미착수,수정중,수정완료,보류"'),
          (COL_RECHK,'"재확인대기,재확인통과,재오픈"')]
    for col,formula in defs:
        dv=DataValidation(type="list",formula1=formula,allow_blank=True)
        ws.add_data_validation(dv)
        L=get_column_letter(col)
        dv.add(f"{L}2:{L}{n+1}")

def add_cf(ws,n):
    Lqc=get_column_letter(COL_QC); rng_qc=f"{Lqc}2:{Lqc}{n+1}"
    ws.conditional_formatting.add(rng_qc,CellIsRule(operator="equal",formula=['"Pass"'],fill=PatternFill("solid",fgColor=FILL_PASS)))
    ws.conditional_formatting.add(rng_qc,CellIsRule(operator="equal",formula=['"Fail"'],fill=PatternFill("solid",fgColor=FILL_FAIL)))
    ws.conditional_formatting.add(rng_qc,CellIsRule(operator="equal",formula=['"Blocked"'],fill=PatternFill("solid",fgColor=FILL_BLOCK)))
    Lfix=get_column_letter(COL_FIX); rng_fix=f"{Lfix}2:{Lfix}{n+1}"
    ws.conditional_formatting.add(rng_fix,CellIsRule(operator="equal",formula=['"수정완료"'],fill=PatternFill("solid",fgColor=FILL_DONE)))
    Lp=get_column_letter(COL_PRIO); rng_p=f"{Lp}2:{Lp}{n+1}"
    ws.conditional_formatting.add(rng_p,CellIsRule(operator="equal",formula=['"P1"'],fill=PatternFill("solid",fgColor=FILL_P1)))

def write_sheet(title,prefix,rows):
    ws=wb.create_sheet(title)
    style_header(ws)
    for i,r in enumerate(rows,1):
        tcid=f"{prefix}-{i:03d}"
        # r=[대분류,중분류,우선순위,케이스명,사전조건,절차,입력,기대결과,비고]
        # 컬럼: 1=담당(공란) / 2=TCID / 3-10=TC정의 / 11-17=QC·결함(공란) / 18=비고
        vals=[""]+[tcid]+r[:8]+[""]*7+[r[8]]
        rr=i+1
        for c,v in enumerate(vals,1):
            cell=ws.cell(rr,c,v)
            cell.font=F(size=9)
            cell.alignment=Alignment(vertical="top",wrap_text=True,
                horizontal="center" if c in (1,COL_TCID,COL_PRIO,COL_QC,COL_SEV,COL_FIX,COL_RECHK) else "left")
            cell.border=BORDER
            if i%2==0:
                if cell.fill.fgColor.rgb in (None,"00000000"):
                    cell.fill=PatternFill("solid",fgColor=FILL_BAND)
    n=len(rows)
    add_dv(ws,n)
    add_cf(ws,n)
    return n,ws

n_rpt,_=write_sheet("TC_결과지","RPT",rpt_rows())
n_api,_=write_sheet("TC_API","API",api_rows())
n_adm,_=write_sheet("TC_어드민","ADM",adm_rows())

# ---- 구현현황 시트 (주요 변경사항 SDS 기준, 개발자 기입용) ----
CHANGES = [
("접수","전형 통과 기준(종합점수/배수) 전환 · BP v2 라디오 제거","프리즘 접수 화면 기능 개발"),
("접수","직무기술서 업로드(RP매칭과 공유)","프리즘 접수 화면 기능 개발"),
("접수","검사요청 확인 화면에 직무기술서 표현","검사 요청 확인 직무기술서 표현"),
("매니저","하드스킬 키워드 관리(문장형식 제거→텍스트형식, 생성/재생성/삭제/추가)","직무기술서 관리 하드스킬 키워드"),
("매니저","BP평가 결과확인 — 한 줄 코멘트 관리","BP평가 결과확인 기능 추가"),
("매니저","BP평가 결과확인 — 자소서 특장점 관리","BP평가 결과확인 기능 추가"),
("매니저","BP평가 결과확인 — 근거요약 관리","BP평가 결과확인 기능 추가"),
("매니저","BP평가 결과확인 — 하드스킬 근거문장 관리","BP평가 결과확인 기능 추가"),
("매니저","BP평가 설정 — 직무기술서 설정(직무별)","BP평가 설정 직무기술서"),
("API","한 줄 코멘트 생성 API (bp-v4-comment)","한 줄 코멘트 생성 API"),
("API","자소서 특장점 요약 API (bp-point-summary)","자소서 특장점 요약 API"),
("API","하드스킬 추출 API (extract-jd-hardskill)","하드스킬 추출 API"),
("API","하드스킬 매칭/검출 API (hardskill-extract-evidence-test)","하드스킬 매칭 API"),
("결과지(PDF)","PDF 전체 family 디자인 적용(레이아웃 현행, 표현만)","전체 라이선스 디자인/종합결과"),
("결과지(PDF)","단일 라이선스 구성(종합결과 없음 + 상단 응시자정보)","전체 라이선스 디자인/종합결과"),
("결과지(PDF)","복수 라이선스 구성(종합결과 페이지)","전체 라이선스 디자인/종합결과"),
("결과지(BP)","BP 콘텐츠(점수·상위n%·전형통과·한줄코멘트·특장점·하드스킬·프로파일·강점검토·문항별·BP상세)","BP평가 콘텐츠 상세"),
("다운로드","프리즘워크 PDF 다운로드 옵션 변경","프리즘워크 PDF 다운로드 옵션"),
("다운로드","엑셀 결과지 변경(전형통과 텍스트·하드스킬 검출)","엑셀파일 결과지 변경"),
]
def build_status_sheet():
    ws=wb.create_sheet("구현현황")
    ws.sheet_view.showGridLines=False
    ws.cell(1,1,"개발 구현현황 — 개발자 기입용 (주요 변경사항 SDS 기준)").font=F(bold=True,size=14,color="1F4E79")
    ws.cell(2,1,"각 기능의 '구현 상태'를 됨/되는중/안됨으로 표시해 주세요. 이 표로 QC의 P1/Blocked를 확정합니다.").font=F(size=10,color="595959")
    hdr=["번호","구분","변경 기능","관련 SDS","구현 상태","예상 완료일","비고"]
    widths=[6,12,54,28,12,13,24]
    hrow=4
    for c,(h,w) in enumerate(zip(hdr,widths),1):
        cell=ws.cell(hrow,c,h); cell.font=F(bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=C_TCDEF); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=BORDER
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[hrow].height=24
    for i,(gu,fn,sds) in enumerate(CHANGES,1):
        r=hrow+i
        for c,v in enumerate([i,gu,fn,sds,"","",""],1):
            cell=ws.cell(r,c,v); cell.font=F(size=10)
            cell.alignment=Alignment(vertical="center",wrap_text=True,horizontal="center" if c in (1,5) else "left"); cell.border=BORDER
    last=hrow+len(CHANGES)
    dv=DataValidation(type="list",formula1='"됨,되는중,안됨,해당없음"',allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"E{hrow+1}:E{last}")
    rngE=f"E{hrow+1}:E{last}"
    ws.conditional_formatting.add(rngE,CellIsRule(operator="equal",formula=['"됨"'],fill=PatternFill("solid",fgColor=FILL_PASS)))
    ws.conditional_formatting.add(rngE,CellIsRule(operator="equal",formula=['"되는중"'],fill=PatternFill("solid",fgColor="FFF2CC")))
    ws.conditional_formatting.add(rngE,CellIsRule(operator="equal",formula=['"안됨"'],fill=PatternFill("solid",fgColor=FILL_FAIL)))
    ws.freeze_panes=f"A{hrow+1}"
build_status_sheet()

# ---- 가이드/대시보드 ----
g=wb["Sheet"]; g.title="가이드"
g.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHI",[3,20,13,11,11,11,11,11,11]): g.column_dimensions[col].width=w
def put(cell,val,**kw):
    c=g[cell]; c.value=val; c.font=F(**{k:v for k,v in kw.items() if k in("bold","size","color","italic")})
    if "fill" in kw: c.fill=PatternFill("solid",fgColor=kw["fill"])
    if "align" in kw: c.alignment=Alignment(horizontal=kw["align"],vertical="center",wrap_text=True)
    return c
put("B2","프리즘 BP 4.2 · QC 테스트케이스",bold=True,size=18,color="1F4E79")
put("B3",f"프리즘 BP(서류평가) 4.2   |   작성일 2026-06-24   |   QC 착수 2026-06-26(금)   |   운영 릴리즈 2026-07-02(목)   |   TC {n_rpt+n_api+n_adm}개(결과지 {n_rpt}·API {n_api}·어드민 {n_adm})",size=10,color="595959")

# ── 테스트 진행 흐름 (큰 그림 먼저) ──
put("B5","■ 테스트는 이렇게 진행됩니다  (먼저 큰 흐름부터)",bold=True,size=13,color="1F4E79")
for rr in (6,7):
    for cc in range(2,10):
        x=g.cell(rr,cc); x.fill=PatternFill("solid",fgColor="DEEBF7"); x.border=BORDER
g.merge_cells("B6:I7")
fc=g["B6"]
fc.value=("①  테스트 실행   →   ②  결과 기록 (Pass / Fail)   →   ③  실패 시 결함·심각도 작성   →   ④  개발 수정   →   ⑤  QA 재확인\n"
          "QA 담당자는  ① · ② · ⑤,   개발자는  ④  를 담당합니다.   ③~⑤ 는 Fail(실패)이 났을 때만 도는 루프입니다.")
fc.font=F(bold=True,size=11,color="041D30")
fc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
g.row_dimensions[6].height=24; g.row_dimensions[7].height=24

# ── 시작하기 (처음이라면) ──
put("B9","■ 시작하기 — 처음이라면 이 3가지만 하면 됩니다",bold=True,size=13,color="1F4E79")
start=["1) 내 담당 시트를 엽니다 — [TC_결과지] · [TC_API] · [TC_어드민] 중 배정받은 영역만 보면 됩니다 (209개를 한 번에 볼 필요 없음).",
       "2) 각 행을 실행한 뒤 [QC결과] 칸에서 Pass / Fail / N/A / Blocked 를 선택하고, [실행일] · [QC담당자] 를 적습니다.",
       "3) Fail 일 때만 [결함 내용] · [심각도] 를 적습니다. 이후 개발이 고치면 다시 확인하는 흐름으로 이어집니다.",
       "※ 1차 점검에서 채울 칸은 [QC결과] · [실행일] · [QC담당자] 3칸뿐입니다. 나머지는 결함이 났을 때만 — 처음부터 모든 칸을 채우려 하지 마세요."]
for i,s in enumerate(start): put(f"B{10+i}",s,size=10)
g["B13"].font=F(size=10,bold=True,color="B5482A")

# ── 참고 · 우선순위 ──
put("B15","■ 참고 · 우선순위 (무엇부터 보나)",bold=True,size=12,color="1F4E79")
pr=[("P1","핵심·필수 — 릴리즈 블로커. 가장 먼저 검증",FILL_P1),
    ("P2","중요 — P1 통과 후 검증",""),
    ("P3","보조·표기 — 시간 여유 시 검증","")]
for i,(k,v,fill) in enumerate(pr):
    put(f"B{16+i}",k,bold=True,size=10,fill=fill or "FFFFFF",align="center"); put(f"C{16+i}",v,size=10)
    g.merge_cells(f"C{16+i}:I{16+i}")

# ── 참고 · 상태 값 ──
put("B20","■ 참고 · 상태 값 (각 칸 드롭다운에서 선택)",bold=True,size=12,color="1F4E79")
leg=[("QC결과","Pass 통과  /  Fail 실패  /  N/A 해당없음  /  Blocked 아직 개발 안 됨→테스트 불가(구현되면 재실행)"),
     ("심각도","치명 릴리즈 불가  /  중 주요 기능 오류  /  경 경미·표기"),
     ("개발 수정상태","미착수  /  수정중  /  수정완료  /  보류   (개발이 기입)"),
     ("재확인(QC)","재확인대기  /  재확인통과  /  재오픈(수정 후에도 실패)")]
for i,(k,v) in enumerate(leg):
    put(f"B{21+i}",k,bold=True,size=10); put(f"C{21+i}",v,size=10); g.merge_cells(f"C{21+i}:I{21+i}")

# ── 참고 · 테스트 방식(액션) ──
put("B26","■ 참고 · 테스트 방식 (결과지 시트 비고의 [ ] 표시 = 무엇을 하는 테스트인지)",bold=True,size=12,color="1F4E79")
act=[("[데이터대조]","매니저 화면 분석값과 결과지 화면의 값이 같은지 눈으로 대조"),
     ("[규칙계산]","정해진 규칙(점수 차이·색 분기 등)대로 나왔는지 확인"),
     ("[정적검사]","화면에 정해진 문구·요소가 있는지 / 없는지 확인"),
     ("[렌더측정]","인쇄·레이아웃이 안 깨지는지(넘침·잘림) 확인"),
     ("[육안]","스크린샷으로 눈으로 확인")]
for i,(k,v) in enumerate(act):
    put(f"B{27+i}",k,bold=True,size=10); put(f"C{27+i}",v,size=10); g.merge_cells(f"C{27+i}:I{27+i}")
put("B32","· API 시트 = API 호출 후 응답(JSON)이 기대대로인지 확인   /   어드민 시트 = 화면에서 클릭·입력 후 동작·표시 결과 확인",size=9,color="595959")

# ── 진행 대시보드 · 시트별 ──
DROW=34
put(f"B{DROW}","■ 진행 대시보드 · 시트별 (자동 집계)",bold=True,size=12,color="1F4E79")
dash_hdr=["시트","전체","미실행","Pass","Fail","N/A","Blocked","통과율"]
for c,h in enumerate(dash_hdr):
    cell=g.cell(DROW+1,2+c,h); cell.font=F(bold=True,color="FFFFFF",size=10)
    cell.fill=PatternFill("solid",fgColor=C_TCDEF); cell.alignment=Alignment(horizontal="center",vertical="center")
sheets=[("TC_결과지",n_rpt),("TC_API",n_api),("TC_어드민",n_adm)]
r0=DROW+2
for i,(sn,nn) in enumerate(sheets):
    r=r0+i
    g.cell(r,2,sn).font=F(size=10,bold=True)
    g.cell(r,3,f"=COUNTA('{sn}'!B2:B2000)")
    g.cell(r,4,f"=C{r}-E{r}-F{r}-G{r}-H{r}")
    g.cell(r,5,f"=COUNTIF('{sn}'!K2:K2000,\"Pass\")")
    g.cell(r,6,f"=COUNTIF('{sn}'!K2:K2000,\"Fail\")")
    g.cell(r,7,f"=COUNTIF('{sn}'!K2:K2000,\"N/A\")")
    g.cell(r,8,f"=COUNTIF('{sn}'!K2:K2000,\"Blocked\")")
    g.cell(r,9,f"=IFERROR(E{r}/(E{r}+F{r}),\"-\")")
    for c in range(2,10):
        cell=g.cell(r,c); cell.alignment=Alignment(horizontal="center"); cell.font=F(size=10)
        if c==9: cell.number_format="0.0%"
rt=r0+len(sheets)
g.cell(rt,2,"합계").font=F(bold=True,size=10)
g.cell(rt,3,f"=SUM(C{r0}:C{rt-1})"); g.cell(rt,4,f"=SUM(D{r0}:D{rt-1})")
g.cell(rt,5,f"=SUM(E{r0}:E{rt-1})"); g.cell(rt,6,f"=SUM(F{r0}:F{rt-1})")
g.cell(rt,7,f"=SUM(G{r0}:G{rt-1})"); g.cell(rt,8,f"=SUM(H{r0}:H{rt-1})")
g.cell(rt,9,f"=IFERROR(E{rt}/(E{rt}+F{rt}),\"-\")")
for c in range(2,10):
    cell=g.cell(rt,c); cell.fill=PatternFill("solid",fgColor="EDEDED"); cell.font=F(bold=True,size=10)
    cell.alignment=Alignment(horizontal="center")
    if c==9: cell.number_format="0.0%"

# ── 진행 대시보드 · 우선순위별 (신규) ──
SNS=("TC_결과지","TC_API","TC_어드민")
rp=rt+2
put(f"B{rp}","■ 진행 대시보드 · 우선순위별 (자동 집계)",bold=True,size=12,color="1F4E79")
for c,h in enumerate(dash_hdr):
    hh="우선순위" if c==0 else h
    cell=g.cell(rp+1,2+c,hh); cell.font=F(bold=True,color="FFFFFF",size=10)
    cell.fill=PatternFill("solid",fgColor=C_TCDEF); cell.alignment=Alignment(horizontal="center",vertical="center")
def _tot(p): return "+".join([f"COUNTIF('{s}'!E2:E2000,\"{p}\")" for s in SNS])
def _cif(p,res): return "+".join([f"COUNTIFS('{s}'!E2:E2000,\"{p}\",'{s}'!K2:K2000,\"{res}\")" for s in SNS])
pri_rows=[("P1",FILL_P1),("P2",""),("P3","")]
pr0=rp+2
for i,(p,fill) in enumerate(pri_rows):
    r=pr0+i
    g.cell(r,2,p).font=F(bold=True,size=10)
    if fill: g.cell(r,2).fill=PatternFill("solid",fgColor=fill)
    g.cell(r,3,f"={_tot(p)}")
    g.cell(r,4,f"=C{r}-E{r}-F{r}-G{r}-H{r}")
    g.cell(r,5,f"={_cif(p,'Pass')}")
    g.cell(r,6,f"={_cif(p,'Fail')}")
    g.cell(r,7,f"={_cif(p,'N/A')}")
    g.cell(r,8,f"={_cif(p,'Blocked')}")
    g.cell(r,9,f"=IFERROR(E{r}/(E{r}+F{r}),\"-\")")
    for c in range(2,10):
        cell=g.cell(r,c); cell.alignment=Alignment(horizontal="center")
        if c!=2: cell.font=F(size=10)
        if c==9: cell.number_format="0.0%"
prt=pr0+len(pri_rows)
put(f"B{prt}","※ P1 통과율이 7/02 릴리즈 가능 여부를 보는 핵심 지표입니다 (P1 전부 Pass + 치명 결함 0 → 릴리즈).",size=9,color="B5482A")

# ── 결함 현황 ──
rd=prt+2
put(f"B{rd}","■ 결함 현황 (자동 집계)",bold=True,size=12,color="B5482A")
for c,h in enumerate(["심각도","건수","","개발 상태","건수"]):
    cell=g.cell(rd+1,2+c,h)
    if h: cell.font=F(bold=True,color="FFFFFF",size=10); cell.fill=PatternFill("solid",fgColor=C_DEFECT)
    cell.alignment=Alignment(horizontal="center")
sev=[("치명","치명"),("중","중"),("경","경")]
fixs=[("미착수","미착수"),("수정중","수정중"),("수정완료","수정완료"),("보류","보류")]
for i,(lab,val) in enumerate(sev):
    g.cell(rd+2+i,2,lab).font=F(size=10)
    parts="+".join([f"COUNTIF('{s}'!O2:O2000,\"{val}\")" for s in SNS])
    g.cell(rd+2+i,3,f"={parts}").alignment=Alignment(horizontal="center")
for i,(lab,val) in enumerate(fixs):
    g.cell(rd+2+i,5,lab).font=F(size=10)
    parts="+".join([f"COUNTIF('{s}'!P2:P2000,\"{val}\")" for s in SNS])
    g.cell(rd+2+i,6,f"={parts}").alignment=Alignment(horizontal="center")

rn=rd+7
put(f"B{rn}","※ 테스트 가능 전제(시트별 우선 확인)",bold=True,size=11,color="595959")
notes=["· [범위] 웹 결과지는 이번 프로젝트 범위 밖 — PDF·엑셀 결과지만 검증(SDS).",
       "· TC_결과지: 샘플 자소서를 접수→분석→결과지까지 돌린 산출물 검증. 값은 매니저 화면 분석값과 대조. 단일/복수 라이선스·엑셀·타 라이선스 페이지(Q~V) 포함.",
       "· TC_API: execute API·공방 프롬프트 배포 후 테스트 가능. 미배포 항목은 Blocked 처리.",
       "· TC_어드민: 어드민 화면 구현분 기준(접수·매니저·설정 포함). 세부 기대값은 화면정의서 확정 시 갱신 필요.",
       "· [확인필요-PM] 하드스킬 입력 방식(파일 업로드 vs plain text)·공방 프롬프트명은 PM 확정 후 반영. 비고 '확인 필요' 행 참조.",
       "· 미구현 기능의 TC는 삭제하지 말고 Blocked로 두어 구현 완료 시 재실행할 것."]
for i,s in enumerate(notes): put(f"B{rn+1+i}",s,size=9,color="595959")

wb.move_sheet("가이드",-(len(wb.sheetnames)-1))
wb.move_sheet("구현현황",-(len(wb.sheetnames)-2))
wb.active=0
import os
out="/Users/gon/Documents/GitHub/HR_report/QC/프리즘BP-4.2-테스트케이스-20260624.xlsx"
wb.save(out)
print("저장:",out)
print(f"결과지 {n_rpt} · API {n_api} · 어드민 {n_adm} · 합계 {n_rpt+n_api+n_adm}")
